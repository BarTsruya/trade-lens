import pandas as pd
import os
import openpyxl

HEBREW_TO_ENGLISH_COLUMNS = {
    "תאריך": 'action_date',
    "סוג פעולה": 'action_type',
    "שם נייר": 'paper_name',
    "מס' נייר / סימבול": 'paper_symbol',
    "כמות": 'quantity',
    "שער ביצוע": 'execution_price',
    "מטבע": 'currency',
    "עמלת פעולה": 'commission_fee',
    "עמלות נלוות": 'additional_fees',
    'תמורה במט"ח': 'total_value_foreign',
    "תמורה בשקלים": 'total_value_shekel',
    "יתרה שקלית": 'shekel_balance',
    "אומדן מס רווחי הון": 'estimated_capital_gains_tax',
}

COLUMN_WIDTH = 20

class StockDataAnalyzer:
    def __init__(self):
        self.df = None
        self.ticker_transactions = {}

    def load_filter_rename_excel(self, file_path):
        self.df = pd.read_excel(file_path)
        self.df = self.df.drop(columns=["מטבע","עמלות נלוות",'תמורה במט"ח','תמורה בשקלים',"יתרה שקלית","שם נייר"])
        self.df.rename(columns=COLUMN_NAMES_DICT, inplace=True)
        self.df['Action Type'] = self.df['Action Type'].replace({
            'קניה חול מטח': 'BUY',
            'מכירה חול מטח': 'SELL'
        })
        filtered_rows = []
        for index, row in self.df.iterrows():
            if row['Action Type'] in ['BUY', 'SELL']:
                filtered_rows.append(row)
        self.df = pd.DataFrame(filtered_rows)
        return self.df

    def add_total_value_column(self):
        if self.df is not None:
            self.df.insert(
                loc=self.df.columns.get_loc('Execution Price')+1,
                column="Total Value",
                value=self.df['Quantity'] * self.df['Execution Price']
            )

    def store_formated_data_to_excel(self, output_path):
        if self.df is not None:
            with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
                self.df.to_excel(writer, index=False, sheet_name='FilteredData')
                
                # Access workbook and worksheet
                workbook  = writer.book
                worksheet = writer.sheets['FilteredData']

                # Set column width (from first to last column)
                # Columns are zero-indexed: here we cover all columns in the DataFrame
                num_cols = len(df.columns)
                worksheet.set_column(0, num_cols - 1, COLUMN_WIDTH)

                # Define format
                center_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
                usd_format = workbook.add_format({'num_format': '$#,##0.00', 'align': 'center', 'valign': 'vcenter'})
                ils_format = workbook.add_format({'num_format': '₪#,##0.00', 'align': 'center', 'valign': 'vcenter'})

                # Apply the format to all data cells including header
                # Get number of rows including header (header is row 0)
                num_rows = len(df) + 1
                

                # Apply format to all cells with data
                worksheet.set_row(0, None, center_format)  # Header row
                for row in range(1, num_rows):
                    worksheet.set_row(row, None, center_format)
                    if 'Execution Price' in self.df.columns:
                        usd_excecution_price_value = self.df.iloc[row-1, self.df.columns.get_loc('Execution Price')]
                        worksheet.write_number(row, self.df.columns.get_loc('Execution Price'), usd_excecution_price_value, usd_format)

                    if "Commission Fee" in self.df.columns:
                        usd_commission_fee_value = self.df.iloc[row-1, self.df.columns.get_loc("Commission Fee")]
                        worksheet.write_number(row, self.df.columns.get_loc("Commission Fee"), usd_commission_fee_value, usd_format)

                    if "Estimated Capital Gains Tax" in self.df.columns:
                        ils_estimated_capital_gains_tax_value = self.df.iloc[row-1, self.df.columns.get_loc("Estimated Capital Gains Tax")]
                        worksheet.write_number(row, self.df.columns.get_loc("Estimated Capital Gains Tax"), ils_estimated_capital_gains_tax_value, ils_format)

                    if "Total Value" in self.df.columns:
                        usd_total_value = self.df.iloc[row-1, self.df.columns.get_loc("Total Value")]
                        worksheet.write_number(row, self.df.columns.get_loc("Total Value"), usd_total_value, usd_format)

            print("Stored formated data to excel successfully!")

    def save_ticker_transactions(self):
        if self.df is not None:
            self.ticker_transactions = {}
            for _, row in self.df.iterrows():
                ticker = row['Ticker']
                if ticker not in self.ticker_transactions:
                    self.ticker_transactions[ticker] = []
                self.ticker_transactions[ticker].append(row)

    def print_remaining_stocks_and_avg_price(self):
        for ticker, transactions in self.ticker_transactions.items():
            qty = 0
            total_cost = 0
            for row in transactions:
                action = row['Action Type']
                quantity = row['Quantity']
                price = row['Execution Price']
                if action == 'BUY':
                    qty += quantity
                    total_cost += quantity * price
                elif action == 'SELL':
                    qty -= quantity
                    total_cost -= min(quantity, qty + quantity) * price  # Remove cost for sold shares
            avg_price = (total_cost / qty) if qty > 0 else 0
            if qty > 0:
                print(f"Ticker: {ticker}, Remaining Shares: {qty}, Average Price: {avg_price:.2f}")

    def create_all_tickers_sheet(self, output_path):
        if self.ticker_transactions:
            import numpy as np
            all_tables = []
            for ticker, transactions in self.ticker_transactions.items():
                # Add a row with the ticker name
                ticker_label_df = pd.DataFrame([[f'Ticker: {ticker}'] + [''] * (len(self.df.columns) - 1)], columns=self.df.columns)
                # Add a header row (will be formatted as bold later)
                header_df = pd.DataFrame([list(self.df.columns)], columns=self.df.columns)
                ticker_df = pd.DataFrame(transactions)
                all_tables.append(ticker_label_df)
                all_tables.append(header_df)
                all_tables.append(ticker_df)
                # Add an empty row for spacing
                empty_row = pd.DataFrame([[''] * len(self.df.columns)], columns=self.df.columns)
                all_tables.append(empty_row)
            combined_df = pd.concat(all_tables, ignore_index=True)
            # Write to Excel and format headings as bold
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                combined_df.to_excel(writer, sheet_name='AllTickers', index=False, header=False)
                workbook = writer.book
                worksheet = writer.sheets['AllTickers']
                bold_font = openpyxl.styles.Font(bold=True)
                center_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                row_idx = 1  # Excel rows are 1-indexed
                for i, row in combined_df.iterrows():
                    # Header rows are those where the row matches the columns exactly
                    if list(row) == list(self.df.columns):
                        for col_idx in range(1, len(self.df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.font = bold_font
                            cell.alignment = center_alignment
                    else:
                        for col_idx in range(1, len(self.df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = center_alignment
                    row_idx += 1
                # Set column widths as defined in COLUMN_WIDTH
                for col_idx in range(1, len(self.df.columns) + 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = COLUMN_WIDTH
            print("Created AllTickers sheet with all ticker tables successfully!")


if __name__ == "__main__":
    stocks_analyzer = StockDataAnalyzer()

    # full database path loaded from IBI   
    file_path = r'C:\Users\user\OneDrive\Documents\Bar\Investments\IBI\analysis\full_list_of_actions_till_june_2025.xlsx'
    
    # Load and filter the data from the Excel file
    df = stocks_analyzer.load_filter_rename_excel(file_path)

    # Add the Total Value column
    stocks_analyzer.add_total_value_column()
    
    # Save ticker transactions
    stocks_analyzer.save_ticker_transactions()
    
    # Print remaining stocks and average price
    stocks_analyzer.print_remaining_stocks_and_avg_price()

    # Store the formatted data to a new Excel file
    output_path = r'C:\Users\user\OneDrive\Documents\Bar\Investments\IBI\analysis\flitered_data3.xlsx'
    stocks_analyzer.store_formated_data_to_excel(output_path)

    # Create AllTickers sheet with all ticker tables
    stocks_analyzer.create_all_tickers_sheet(output_path)