import pandas as pd
import os
import openpyxl

COLUMN_NAMES_DICT = {
    "תאריך": 'Date',
    "סוג פעולה": 'Action Type',
    "שם נייר": 'Security Name',
    "מס' נייר / סימבול": 'Ticker',
    "כמות": 'Quantity',
    "שער ביצוע": 'Execution Price',
    "עמלת פעולה": 'Commission Fee',
    "אומדן מס רווחי הון": 'Estimated Capital Gains Tax'
}

COLUMN_WIDTH = 20


def load_filter_rename_excel(file_path):
    # Read the excel file
    df = pd.read_excel(file_path)

    # Drop unnecessary columns
    df = df.drop(columns=["מטבע","עמלות נלוות",'תמורה במט"ח','תמורה בשקלים',"יתרה שקלית","שם נייר"])

    # Rename columns for clarity
    df.rename(columns=COLUMN_NAMES_DICT, inplace=True)

    # Rename columns for easier access
    df['Action Type'] = df['Action Type'].replace({
        'קניה חול מטח': 'BUY',
        'מכירה חול מטח': 'SELL'
    })

    # Create a new DataFrame to store the only sell and buy actions
    filtered_rows = []

    for index, row in df.iterrows():
        if row['Action Type'] in ['BUY', 'SELL']:
            filtered_rows.append(row)

    df_filtered = pd.DataFrame(filtered_rows)
    return df_filtered


def add_total_value_column(df):
    # Create a new column for the sum of shares
    df.insert(loc=df.columns.get_loc('Execution Price')+1,column="Total Value", value=df['Quantity'] * df['Execution Price'])



def store_formated_data_to_excel(df, output_path):
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='FilteredData')
        
        # Access workbook and worksheet
        workbook  = writer.book
        worksheet = writer.sheets['FilteredData']

        # Set column width (from first to last column)
        # Columns are zero-indexed: here we cover all columns in the DataFrame
        num_cols = len(df.columns)
        worksheet.set_column(0, num_cols - 1, COLUMN_WIDTH)

        # Define format
        center_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        usd_format = workbook.add_format({'num_format': '$#,##0.00', 'align': 'center', 'valign': 'vcenter'})
        ils_format = workbook.add_format({'num_format': '₪#,##0.00', 'align': 'center', 'valign': 'vcenter'})

        # Apply the format to all data cells including header
        # Get number of rows including header (header is row 0)
        num_rows = len(df) + 1
        

        # Apply format to all cells with data
        worksheet.set_row(0, None, center_format)  # Header row
        for row in range(1, num_rows):
            worksheet.set_row(row, None, center_format)
            if 'Execution Price' in df.columns:
                usd_excecution_price_value = df.iloc[row-1, df.columns.get_loc('Execution Price')]
                worksheet.write_number(row, df.columns.get_loc('Execution Price'), usd_excecution_price_value, usd_format)

            if "Commission Fee" in df.columns:
                usd_commission_fee_value = df.iloc[row-1, df.columns.get_loc("Commission Fee")]
                worksheet.write_number(row, df.columns.get_loc("Commission Fee"), usd_commission_fee_value, usd_format)

            if "Estimated Capital Gains Tax" in df.columns:
                ils_estimated_capital_gains_tax_value = df.iloc[row-1, df.columns.get_loc("Estimated Capital Gains Tax")]
                worksheet.write_number(row, df.columns.get_loc("Estimated Capital Gains Tax"), ils_estimated_capital_gains_tax_value, ils_format)

            if "Total Value" in df.columns:
                usd_total_value = df.iloc[row-1, df.columns.get_loc("Total Value")]
                worksheet.write_number(row, df.columns.get_loc("Total Value"), usd_total_value, usd_format)

    print("Stored formated data to excel successfully!")


def get_ticker_list(df):
    # Get the unique tickers from the DataFrame
    return df['Ticker'].unique().tolist()


def create_ticker_actions_table(df, ticker):
    """
    Returns a DataFrame containing all actions for the given ticker.
    """
    return df[df['Ticker'] == ticker].reset_index(drop=True)


def store_ticker_tables_to_excel(df, output_path):
    ticker_list = get_ticker_list(df)
    # Check if file exists
    file_exists = os.path.exists(output_path)
    if file_exists:
        # Load existing file and add a new sheet
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            start_row = 0
            worksheet = writer.book.create_sheet('AllTickers') if 'AllTickers' not in writer.book.sheetnames else writer.book['AllTickers']
            for ticker in ticker_list:
                ticker_df = create_ticker_actions_table(df, ticker)
                if not ticker_df.empty:
                    worksheet.cell(row=start_row+1, column=1, value=f"Ticker: {ticker}")
                    start_row += 1
                    for col_num, col_name in enumerate(ticker_df.columns, 1):
                        worksheet.cell(row=start_row+1, column=col_num, value=col_name)
                    for row_num, row in enumerate(ticker_df.values, start=start_row+2):
                        for col_num, value in enumerate(row, 1):
                            worksheet.cell(row=row_num, column=col_num, value=value)
                    start_row += len(ticker_df) + 2
            # Set column widths
            for col in range(1, len(df.columns) + 1):
                col_letter = worksheet.cell(row=1, column=col).column_letter
                worksheet.column_dimensions[col_letter].width = COLUMN_WIDTH
        print(f"Added AllTickers sheet to {output_path} successfully!")
    else:
        # Create new file with AllTickers sheet
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            start_row = 0
            worksheet = writer.book.add_worksheet('AllTickers')
            writer.sheets['AllTickers'] = worksheet
            # Set column widths
            worksheet.set_column(0, len(df.columns) - 1, COLUMN_WIDTH)
            for ticker in ticker_list:
                ticker_df = create_ticker_actions_table(df, ticker)
                if not ticker_df.empty:
                    worksheet.write(start_row, 0, f"Ticker: {ticker}")
                    start_row += 1
                    for col_num, col_name in enumerate(ticker_df.columns):
                        worksheet.write(start_row, col_num, col_name)
                    for row in ticker_df.itertuples(index=False):
                        start_row += 1
                        for col_num, value in enumerate(row):
                            worksheet.write(start_row, col_num, value)
                    start_row += 2  # Blank row after each ticker
        print(f"Created new file with AllTickers sheet at {output_path} successfully!")


if __name__ == "__main__":
    # full database path loaded from IBI   
    file_path = r'C:\Users\user\OneDrive\Documents\Bar\Investments\IBI\analysis\full_list_of_actions_till_june_2025.xlsx'

    # Load the data from the Excel file
    df = load_filter_rename_excel(file_path)

    # Add a new column for the sum of shares
    add_total_value_column(df)

    # Store the formatted data to a new Excel file
    output_path = r'C:\Users\user\OneDrive\Documents\Bar\Investments\IBI\analysis\flitered_data.xlsx'
    store_formated_data_to_excel(df, output_path)

    # Store the ticker tables to a new Excel file
    ticker_output_path = r'C:\Users\user\OneDrive\Documents\Bar\Investments\IBI\analysis\flitered_data.xlsx'
    store_ticker_tables_to_excel(df, ticker_output_path)